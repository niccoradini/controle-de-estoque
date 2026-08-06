import collections
import json
import re
import sys

import openpyxl


workbook = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
result = []

for worksheet in workbook.worksheets:
    records = []
    for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_number == 1 or not any(value is not None for value in values):
            continue
        product, manufacturer, model, good, defective, table_date, flag_good, flag_defective = values[1:9]
        if not model:
            continue
        match = re.match(r"^(ALLIED|ASSURANT|TROCAFONE)\s+(.+)$", str(model).strip(), re.I)
        source = match.group(1).upper() if match else "OTHER"
        display_name = match.group(2).strip() if match else str(model).strip()
        records.append(
            {
                "row": row_number,
                "source": source,
                "product": product,
                "manufacturer": manufacturer,
                "model": display_name,
                "good": good,
                "defective": defective,
                "table_date": table_date,
                "flag_good": flag_good,
                "flag_defective": flag_defective,
            }
        )

    by_source = collections.Counter(record["source"] for record in records)
    by_source_date = collections.Counter(
        (record["source"], str(record["table_date"])[:10]) for record in records
    )
    assurant = [record for record in records if record["source"] == "ASSURANT"]
    result.append(
        {
            "sheet": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "record_count": len(records),
            "by_source": dict(sorted(by_source.items())),
            "by_source_date": [
                {"source": source, "date": date, "count": count}
                for (source, date), count in sorted(by_source_date.items())
            ],
            "assurant_count": len(assurant),
            "assurant_manufacturers": dict(
                sorted(collections.Counter(record["manufacturer"] for record in assurant).items())
            ),
            "assurant_bad_value_order": [
                record for record in assurant if float(record["good"] or 0) < float(record["defective"] or 0)
            ],
            "assurant_iphone_14_128": [
                record for record in assurant if "IPHONE 14 128GB" in record["model"].upper()
            ],
            "assurant_first": assurant[:8],
            "assurant_last": assurant[-8:],
        }
    )

print(json.dumps(result, ensure_ascii=False, default=str, indent=2))
