#!/usr/bin/env python3
import json
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

REQUIRED = ['Material', 'Denominação', 'Nº de série', 'Centro', 'Depósito', 'Tipo de estoque', 'Status sistema', 'Modificado por', 'Modificado em']
STORE_NAMES = {
    'ESTOQUE LOJA BQ LUCAS.xlsx': 'BQ Lucas',
    'ESTOQUE LOJA PATIO.xlsx': 'Pátio',
    'ESTOQUE LOJA AVENIDA.xlsx': 'Avenida',
    'ESTOQUE 89MN.xlsx': 'BQ Lucas',
    'ESTOQUE 283H.xlsx': 'Pátio',
    'ESTOQUE 210H.xlsx': 'Avenida',
}

if len(sys.argv) < 2:
    raise SystemExit('Informe uma ou mais planilhas de estoque.')

stores = []
snapshot_date = '2026-09-01'
filenames = [value for value in sys.argv[1:] if not value.startswith('--snapshot-date=')]
for value in sys.argv[1:]:
    if value.startswith('--snapshot-date='):
        snapshot_date = value.split('=', 1)[1]

for filename in filenames:
    path = Path(filename).resolve()
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(value or '').strip() for value in next(sheet.iter_rows(values_only=True))]
    if not all(header in headers for header in REQUIRED):
        raise SystemExit(f'Cabeçalhos inesperados em {path.name}: {headers!r}')
    header_index = {header: index for index, header in enumerate(headers)}
    grouped = defaultdict(lambda: {'available': 0, 'incoming': 0, 'repair': 0, 'ignored': 0, 'latestModifiedOn': ''})
    centers = set()
    serials = set()
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        material = values[header_index['Material']]
        technical_name = values[header_index['Denominação']]
        serial = values[header_index['Nº de série']]
        center = values[header_index['Centro']]
        deposit = values[header_index['Depósito']]
        stock_type = values[header_index['Tipo de estoque']]
        status = values[header_index['Status sistema']]
        modified_by = values[header_index['Modificado por']]
        modified_value = values[header_index['Modificado em']]
        material = str(material or '').strip()
        technical_name = str(technical_name or '').strip()
        serial = str(serial or '').strip()
        center = str(center or '').strip()
        deposit = str(deposit or '').strip().upper()
        status = str(status or '').strip().upper()
        if not material or not technical_name or not serial or not center:
            raise SystemExit(f'Linha {source_row} incompleta em {path.name}.')
        if serial.casefold() in serials:
            raise SystemExit(f'Série duplicada em {path.name}: {serial}')
        serials.add(serial.casefold())
        centers.add(center)
        if isinstance(modified_value, datetime):
            modified_on = modified_value.date().isoformat()
        elif isinstance(modified_value, date):
            modified_on = modified_value.isoformat()
        else:
            modified_on = str(modified_value or '').strip()[:10]
        item = grouped[(material, technical_name)]
        item['latestModifiedOn'] = max(item['latestModifiedOn'], modified_on)
        if deposit == 'RPAR':
            item['repair'] += 1
        elif status == 'DEPS':
            item['available'] += 1
        elif status == 'DEPS NREM':
            item['incoming'] += 1
        else:
            item['ignored'] += 1
    if len(centers) != 1:
        raise SystemExit(f'A planilha {path.name} possui mais de um centro: {sorted(centers)}')
    store_name = STORE_NAMES.get(path.name, path.stem.replace('ESTOQUE LOJA ', '').title())
    code = store_name.lower().replace('á', 'a').replace('ã', 'a').replace(' ', '-')
    items = [
        {'materialCode': material, 'technicalName': technical_name, **counts}
        for (material, technical_name), counts in sorted(grouped.items())
    ]
    stores.append({
        'code': code,
        'name': store_name,
        'center': next(iter(centers)),
        'sourceFile': path.name,
        'snapshotDate': snapshot_date,
        'sourceRows': len(serials),
        'items': items,
    })

output = Path(f'data/network-inventory-{snapshot_date}.json')
output.write_text(json.dumps({'snapshotDate': snapshot_date, 'stores': stores}, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({store['name']: {'rows': store['sourceRows'], 'materials': len(store['items'])} for store in stores}, ensure_ascii=False))
