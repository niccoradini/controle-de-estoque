#!/usr/bin/env python3
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


if len(sys.argv) != 2:
    raise SystemExit('Uso: python scripts/parse-stock-report-2026-08-25.py caminho/ESTOQUE25.08.xlsx')

source_path = Path(sys.argv[1]).resolve()
workbook = load_workbook(source_path, read_only=True, data_only=True)
sheet = workbook.active
headers = [str(value or '').strip() for value in next(sheet.iter_rows(values_only=True))]
expected_headers = ['Material', 'Denominação', 'Nº de série', 'Centro', 'Depósito']
extended_headers = expected_headers + ['Tipo de estoque', 'Status sistema', 'Modificado por', 'Modificado em']
if headers not in (expected_headers, extended_headers):
    raise SystemExit(f'Cabeçalhos inesperados: {headers!r}')

available_rows = []
incoming_rows = []
excluded_rpar = 0
repair_rows = []
status_summary = Counter()
excluded_status_rows = 0
serials = set()

for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    normalized = [str(value or '').strip() for value in values]
    material, technical_name, serial_number, center, deposit = normalized[:5]
    stock_type = normalized[5] if len(normalized) > 5 else '01'
    reported_status = normalized[6].upper() if len(normalized) > 6 else ''
    if not material or not technical_name or not serial_number or not center:
        raise SystemExit(f'Linha {source_row} incompleta.')
    serial_key = serial_number.casefold()
    if serial_key in serials:
        raise SystemExit(f'Número de série duplicado: {serial_number}')
    serials.add(serial_key)

    deposit = deposit.upper()
    if deposit == 'RPAR':
        excluded_rpar += 1
        repair_rows.append({
            'sourceRow': source_row,
            'material': material,
            'technicalName': technical_name,
            'serialNumber': serial_number,
            'center': center,
            'deposit': 'RPAR',
        })
        continue

    system_status = reported_status or ('DEPS NREM' if not deposit else 'DEPS')
    if system_status in ('LIDI', 'LIDI NREM'):
        excluded_status_rows += 1
        status_summary[system_status] += 1
        continue
    if system_status not in ('DEPS', 'DEPS NREM'):
        raise SystemExit(f'Status desconhecido na linha {source_row}: {system_status!r}')
    row = {
        'sourceRow': source_row,
        'material': material,
        'technicalName': technical_name,
        'serialNumber': serial_number,
        'center': center,
        'deposit': deposit or 'NREM',
        'stockType': stock_type,
        'systemStatus': system_status,
    }
    (incoming_rows if system_status == 'DEPS NREM' else available_rows).append(row)
    status_summary[system_status] += 1

material_count = len({row['material'] for row in available_rows + incoming_rows})
result = {
    'source': source_path.name,
    'importedAt': '2026-08-25',
    'migrationNumber': 47,
    'headers': headers,
    'sourceRowCount': len(serials),
    'availableDeposits': ['EXPO', 'LOJA', 'LVUT'],
    'incomingDeposits': ['DEPS', 'NREM'],
    'excludedDeposits': ['RPAR'],
    'availableStatuses': ['DEPS'],
    'incomingStatuses': ['DEPS NREM'],
    'excludedStatuses': ['LIDI', 'LIDI NREM'],
    'normalizedBlankDeposit': 'NREM',
    'normalizedBlankDepositCount': len(incoming_rows),
    'statusSummary': dict(status_summary),
    'excludedRowCount': excluded_rpar,
    'excludedStatusRowCount': excluded_status_rows,
    'expectedExcludedRowCount': excluded_rpar,
    'expectedProductCount': material_count,
    'expectedAvailableQuantity': len(available_rows),
    'incomingRowCount': len(incoming_rows),
    'rows': available_rows,
    'incomingRows': incoming_rows,
    'repairRows': repair_rows,
}

output_path = Path('data/basis-serial-stock-2026-08-25-excluding-rpar-with-incoming.json')
output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'sourceRows': len(serials),
    'materials': material_count,
    'available': len(available_rows),
    'incoming': len(incoming_rows),
    'excludedRpar': excluded_rpar,
}, ensure_ascii=False))
