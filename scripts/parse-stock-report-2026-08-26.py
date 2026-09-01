#!/usr/bin/env python3
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


if len(sys.argv) not in (2, 3):
    raise SystemExit('Uso: python scripts/parse-stock-report-2026-08-26.py caminho/ESTOQUEdd.mm.xlsx [numero_migracao]')

source_path = Path(sys.argv[1]).resolve()
migration_number = int(sys.argv[2]) if len(sys.argv) == 3 else 58
date_match = re.search(r'(\d{2})\.(\d{2})', source_path.name)
if not date_match:
    raise SystemExit('O nome do arquivo precisa conter a data no formato dd.mm.')
snapshot_date = f'2026-{date_match.group(2)}-{date_match.group(1)}'
workbook = load_workbook(source_path, read_only=True, data_only=True)
sheet = workbook.active
headers = [str(value or '').strip() for value in next(sheet.iter_rows(values_only=True))]
legacy_headers = ['Material', 'Denominação', 'Nº de série', 'Centro', 'Depósito', 'Modificado por', 'Modificado em']
required_headers = ['Material', 'Denominação', 'Nº de série', 'Centro', 'Depósito']
if not all(header in headers for header in required_headers):
    raise SystemExit(f'Cabeçalhos inesperados: {headers!r}')
header_index = {header: index for index, header in enumerate(headers)}
is_current = 'Status sistema' in header_index

available_rows = []
incoming_rows = []
repair_rows = []
excluded_status_rows = []
status_summary = Counter()
serials = set()

for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    material = values[header_index['Material']]
    technical_name = values[header_index['Denominação']]
    serial_number = values[header_index['Nº de série']]
    center = values[header_index['Centro']]
    deposit = values[header_index['Depósito']]
    stock_type = values[header_index['Tipo de estoque']] if 'Tipo de estoque' in header_index else '01'
    system_status = values[header_index['Status sistema']] if is_current else ('DEPS NREM' if not deposit else 'DEPS')
    modified_by = values[header_index['Modificado por']] if 'Modificado por' in header_index else ''
    modified_value = values[header_index['Modificado em']] if 'Modificado em' in header_index else None
    material = str(material or '').strip()
    technical_name = str(technical_name or '').strip()
    serial_number = str(serial_number or '').strip()
    center = str(center or '').strip()
    deposit = str(deposit or '').strip().upper()
    stock_type = str(stock_type or '').strip()
    system_status = str(system_status or '').strip().upper()
    modified_by = str(modified_by or '').strip()
    if isinstance(modified_value, datetime):
        modified_on = modified_value.date().isoformat()
    elif isinstance(modified_value, date):
        modified_on = modified_value.isoformat()
    else:
        modified_on = str(modified_value or '').strip()[:10]
    if not material or not technical_name or not serial_number or not center:
        raise SystemExit(f'Linha {source_row} incompleta.')
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', modified_on):
        raise SystemExit(f'Data inválida na linha {source_row}: {modified_on!r}')
    serial_key = serial_number.casefold()
    if serial_key in serials:
        raise SystemExit(f'Número de série duplicado: {serial_number}')
    serials.add(serial_key)

    if deposit == 'RPAR':
        repair_rows.append({
            'sourceRow': source_row,
            'material': material,
            'technicalName': technical_name,
            'serialNumber': serial_number,
            'center': center,
            'deposit': 'RPAR',
            'modifiedBy': modified_by,
            'modifiedOn': modified_on,
        })
        continue

    row = {
        'sourceRow': source_row,
        'material': material,
        'technicalName': technical_name,
        'serialNumber': serial_number,
        'center': center,
        'deposit': deposit or 'NREM',
        'stockType': stock_type,
        'systemStatus': system_status,
        'modifiedBy': modified_by,
        'modifiedOn': modified_on,
    }
    if system_status == 'DEPS':
        available_rows.append(row)
    elif system_status == 'DEPS NREM':
        incoming_rows.append(row)
    else:
        excluded_status_rows.append(row)
    status_summary[system_status] += 1

material_count = len({row['material'] for row in available_rows + incoming_rows})
result = {
    'source': source_path.name,
    'importedAt': snapshot_date,
    'migrationNumber': migration_number,
    'headers': headers,
    'sourceRowCount': len(serials),
    'availableDeposits': ['EXPO', 'LOJA', 'LVUT'],
    'incomingDeposits': ['DEPS', 'NREM'],
    'excludedDeposits': ['RPAR'],
    'availableStatuses': ['DEPS'],
    'incomingStatuses': ['DEPS NREM'],
    'excludedStatuses': sorted({row['systemStatus'] for row in excluded_status_rows}),
    'normalizedBlankDeposit': 'NREM',
    'normalizedBlankDepositCount': len(incoming_rows),
    'statusSummary': dict(status_summary),
    'excludedRowCount': len(repair_rows),
    'excludedStatusRowCount': len(excluded_status_rows),
    'expectedExcludedRowCount': len(repair_rows),
    'expectedProductCount': material_count,
    'expectedAvailableQuantity': len(available_rows),
    'incomingRowCount': len(incoming_rows),
    'rows': available_rows,
    'incomingRows': incoming_rows,
    'repairRows': repair_rows,
}

output_path = Path(f'data/basis-serial-stock-{snapshot_date}-excluding-rpar-with-incoming.json')
output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'sourceRows': len(serials),
    'materials': material_count,
    'available': len(available_rows),
    'incoming': len(incoming_rows),
    'excludedRpar': len(repair_rows),
}, ensure_ascii=False))
